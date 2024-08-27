import pandas as pd
from openpyxl import load_workbook


def convert_date_and_save(
    filepath, sheet_index, date_col_index, output_filename, date_format="%d-%m-%Y"
):
    """
    This function reads an Excel file, converts a specific column in a specified sheet to date format,
    and saves the modified data as a new Excel file.

    Args:
        filepath (str): Path to the Excel file to read.
        sheet_index (int): Index of the sheet to modify (0-based indexing).
        date_col_index (int): Index of the column containing date strings (0-based indexing).
        output_filename (str): Name of the output Excel file to save.
        date_format (str, optional): Format of the date strings in the column. Defaults to '%d-%m-%Y'.
    """
    # Load the workbook
    workbook = load_workbook(filename=filepath, data_only=True)

    # Access the sheet to modify
    sheet = workbook[workbook.sheetnames[sheet_index]]

    # Convert the column to date
    for row in sheet.iter_rows(min_row=2):  # Skip header row (assuming row 1)
        try:
            row[date_col_index].value = pd.to_datetime(
                row[date_col_index].value, format=date_format, errors="coerce"
            )
        except (ValueError, AttributeError):  # Handle errors during conversion
            print(f"Error converting cell {row[date_col_index].coordinate} to date.")

    # Save the modified workbook with a new filename
    workbook.save(filename=output_filename)


# Replace these values with your actual data
filepath = r"/mnt/e/work/nuevas-asignaciones-de-clientes/v2/clients-table.xlsx"
sheet_index = 2  # Third sheet (0-based indexing)
date_col_index = 6  # Eighth column (0-based indexing)
date_format = "%d-%m-%Y"  # Adjust if your date format is different
output_filename = "/mnt/e/work/nuevas-asignaciones-de-clientes/v2/clients_table_modified_dates.xlsx"  # New output filename

convert_date_and_save(
    filepath, sheet_index, date_col_index, output_filename, date_format
)
